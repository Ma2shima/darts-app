"""
Darts Improvement Tracker / ダーツ上達確認アプリ
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import math
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ============================================================
# ダーツ盤定数（WDF 標準比率）
# ============================================================
NUMBERS = [20, 1, 18, 4, 13, 6, 10, 15, 2, 17, 3, 19, 7, 16, 8, 11, 14, 9, 12, 5]

CANVAS_SIZE  = 500
CX = CY = 250            # キャンバス中心

R_SURROUND   = 238       # 外縁ブラックリング外周
R_DBL_OUT    = 212       # ダブルリング外周 = 採点エリア外周 (170mm)
R_DBL_IN     = 202       # ダブルリング内周 (162mm)
R_TRIPLE_OUT = 133       # トリプルリング外周 (107mm)
R_TRIPLE_IN  = 124       # トリプルリング内周 (99mm)
R_OUTER_BULL = 20        # アウターブル半径 (15.9mm)
R_BULL_EYE   = 8         # ブルズアイ半径 (6.35mm)

# mm 換算スケール: R_DBL_OUT px = 170mm
PX_TO_MM = 170.0 / R_DBL_OUT
MISS_DISTANCE_MM = 300   # ミス時の固定距離(mm)

# ダーツ盤カラー
SEG_DARK   = "#1a1a1a"   # シングルA（黒）
SEG_LIGHT  = "#f0ede0"   # シングルB（クリーム）
RING_RED   = "#cc2200"   # リングA（赤）
RING_BLUE  = "#1144cc"   # リングB（青）
SURROUND_C = "#151515"   # 外縁リング
WIRE_C     = "#303030"   # ワイヤー線
BULL_GREEN = "#005500"   # アウターブル（グリーン）

# ボタンカラー
BTN_BLUE     = "#2980b9"
BTN_DARK     = "#1a6fa8"
BTN_DISABLED = "#7f8c8d"


# ============================================================
# 言語文字列
# ============================================================
STRINGS = {
    "ja": {
        "title":            "ダーツ上達確認アプリ",
        "lang_btn":         "EN",
        "name_label":       "体験者名：",
        "reset_btn":        "リセット",
        "status_frame":     "現在の状態",
        "set_label":        "セット {n}",
        "throw_prompt":     "{n}投目を投げてください",
        "instruction":      "ダーツ盤をクリックして\n着弾点を入力\n（盤外クリック＝ミス）",
        "set1_frame":       "セット1 記録",
        "set2_frame":       "セット2 記録",
        "throw_recorded":   "  {n}投目：({x:+.0f},{y:+.0f})mm ✎",
        "throw_miss":       "  {n}投目：ミス（盤外） ✎",
        "throw_pending":    "  {n}投目：未記録",
        "result_frame":     "結果",
        "next_set_btn":     "セット2へ進む",
        "save_btn":         "Excelに保存",
        "canvas_hint":      "クリックで着弾点を記録（盤外＝ミス）",
        "centroid_label":   "S{n}重心",
        "set1_complete_title": "セット1完了",
        "set1_complete_msg":   "セット1の記録が完了しました。\n「セット2へ進む」ボタンを押してください。",
        "set1_done_status":    "セット1完了！\n「セット2へ進む」を押してください",
        "set2_done_status":    "セット2完了！\n結果を確認してください",
        "set2_instruction":    "オレンジ色のマーカーで\n記録されます",
        "edit_mode_msg":       "✎ セット{s}：{n}投目を修正中\n盤をクリックして再入力",
        "edit_cancel_tip":     "（同じ行を再クリックでキャンセル）",
        "reset_confirm_title": "確認",
        "reset_confirm_msg":   "全データをリセットしますか？",
        "no_excel_title":      "openpyxlが必要です",
        "no_excel_msg":        "pip install openpyxl を実行してください。",
        "save_done_title":     "保存完了・リセット",
        "save_done_msg":       "Excelを保存しました。\n{path}\n\nデータをリセットしました。",
        "excel_filetypes":     "Excel ファイル",
        "result_spread":       "■ バラつき（グルーピング）",
        "result_accuracy":     "■ 正確さ（中心距離平均）",
        "result_round1":       "  1回目：{v:.1f} mm",
        "result_round2":       "  2回目：{v:.1f} mm",
        "result_improved":     "  改善率：{v:.0f}% ↑ 上達!",
        "result_worsened":     "  変化率：{v:.0f}% (悪化)",
        "popup_title_good":    "結果：上達しました！",
        "popup_title_bad":     "結果",
        "popup_body": (
            "【結果】\n\n"
            "バラつき（グルーピング半径）\n"
            "  1回目：{r1:.1f} mm\n"
            "  2回目：{r2:.1f} mm\n"
            "  改善率：{imp:.0f}%\n\n"
            "正確さ（ブル中心からの平均距離）\n"
            "  1回目：{d1:.1f} mm\n"
            "  2回目：{d2:.1f} mm\n"
            "  改善率：{acc:.0f}%"
        ),
        "xl_sheet": "ダーツ記録", "xl_main_title": "ダーツ体験記録",
        "xl_name": "体験者名",   "xl_date": "記録日時",
        "xl_set1": "■ セット1 記録", "xl_set2": "■ セット2 記録",
        "xl_result": "■ 比較結果",   "xl_throw": "投擲",
        "xl_x": "X座標 (mm)",        "xl_y": "Y座標 (mm)",
        "xl_dist": "ブルからの距離 (mm)", "xl_group_r": "グルーピング半径",
        "xl_mm": "mm",               "xl_item": "項目",
        "xl_round1": "1回目",        "xl_round2": "2回目",
        "xl_improve": "改善率",      "xl_judge": "判定",
        "xl_spread": "バラつき（グルーピング半径）",
        "xl_accuracy": "正確さ（中心距離平均）",
        "xl_good": "上達",           "xl_bad": "変化なし/悪化",
        "xl_comment": "コメント",    "xl_throw_n": "{n}投目",
        "xl_miss": "ミス（盤外）",   "miss_marker": "MISS",
    },
    "en": {
        "title":            "Darts Improvement Tracker",
        "lang_btn":         "日本語",
        "name_label":       "Name:",
        "reset_btn":        "Reset",
        "status_frame":     "Status",
        "set_label":        "Round {n}",
        "throw_prompt":     "Throw #{n}",
        "instruction":      "Click board to record\nlanding point\n(Outside = Miss)",
        "set1_frame":       "Round 1 Record",
        "set2_frame":       "Round 2 Record",
        "throw_recorded":   "  #{n}: ({x:+.0f},{y:+.0f})mm ✎",
        "throw_miss":       "  #{n}: Miss (outside) ✎",
        "throw_pending":    "  #{n}: --",
        "result_frame":     "Results",
        "next_set_btn":     "Go to Round 2",
        "save_btn":         "Export to Excel",
        "canvas_hint":      "Click board to record  (outside = miss)",
        "centroid_label":   "R{n} center",
        "set1_complete_title": "Round 1 Complete",
        "set1_complete_msg":   "Round 1 recorded.\nPress 'Go to Round 2'.",
        "set1_done_status":    "Round 1 done!\nPress 'Go to Round 2'",
        "set2_done_status":    "Round 2 done!\nCheck results",
        "set2_instruction":    "Orange markers will be\nused for Round 2",
        "edit_mode_msg":       "✎ Editing Round {s}, Throw {n}\nClick board to re-enter",
        "edit_cancel_tip":     "(Click same row to cancel)",
        "reset_confirm_title": "Confirm",
        "reset_confirm_msg":   "Reset all data?",
        "no_excel_title":      "openpyxl required",
        "no_excel_msg":        "Run: pip install openpyxl",
        "save_done_title":     "Saved & Reset",
        "save_done_msg":       "Excel file saved.\n{path}\n\nData has been reset.",
        "excel_filetypes":     "Excel files",
        "result_spread":       "■ Grouping (Spread)",
        "result_accuracy":     "■ Accuracy (Avg center dist)",
        "result_round1":       "  Round 1: {v:.1f} mm",
        "result_round2":       "  Round 2: {v:.1f} mm",
        "result_improved":     "  Improvement: {v:.0f}% ↑ Better!",
        "result_worsened":     "  Change: {v:.0f}% (worse)",
        "popup_title_good":    "Result: Improved!",
        "popup_title_bad":     "Result",
        "popup_body": (
            "[Result]\n\n"
            "Grouping radius (spread)\n"
            "  Round 1: {r1:.1f} mm\n"
            "  Round 2: {r2:.1f} mm\n"
            "  Improvement: {imp:.0f}%\n\n"
            "Accuracy (avg distance from bull)\n"
            "  Round 1: {d1:.1f} mm\n"
            "  Round 2: {d2:.1f} mm\n"
            "  Improvement: {acc:.0f}%"
        ),
        "xl_sheet": "Darts Record",       "xl_main_title": "Darts Experience Record",
        "xl_name": "Name",                 "xl_date": "Date/Time",
        "xl_set1": "■ Round 1 Record",    "xl_set2": "■ Round 2 Record",
        "xl_result": "■ Comparison",      "xl_throw": "Throw",
        "xl_x": "X coord (mm)",            "xl_y": "Y coord (mm)",
        "xl_dist": "Distance from bull (mm)", "xl_group_r": "Grouping radius",
        "xl_mm": "mm",                     "xl_item": "Item",
        "xl_round1": "Round 1",            "xl_round2": "Round 2",
        "xl_improve": "Improvement",       "xl_judge": "Judgment",
        "xl_spread": "Grouping radius (spread)",
        "xl_accuracy": "Accuracy (avg center dist)",
        "xl_good": "Improved",             "xl_bad": "No change / Worse",
        "xl_comment": "Comment",           "xl_throw_n": "Throw {n}",
        "xl_miss": "Miss (outside board)", "miss_marker": "MISS",
    },
}


# ============================================================
# macOS 対応ブルーボタン（Label ベース）
# macOS の Aqua テーマは tk.Button の bg を無視するため Label で代替
# ============================================================
class BlueBtn(tk.Label):
    def __init__(self, parent, text="", command=None, font=None, **kwargs):
        super().__init__(
            parent, text=text,
            bg=BTN_BLUE, fg="white",
            font=font or ("Helvetica", 11),
            cursor="hand2", padx=8, pady=7,
            relief="flat", **kwargs
        )
        self._cmd = command
        self._on  = True
        self.bind("<Button-1>", self._click)
        self.bind("<Enter>",    lambda e: self._hover(True))
        self.bind("<Leave>",    lambda e: self._hover(False))

    def _click(self, _):
        if self._on and self._cmd:
            self._cmd()

    def _hover(self, enter):
        if self._on:
            self.config(bg=BTN_DARK if enter else BTN_BLUE)

    def set_state(self, state):
        self._on = (state == "normal")
        self.config(
            bg=BTN_BLUE if self._on else BTN_DISABLED,
            cursor="hand2" if self._on else "arrow"
        )


# ============================================================
# アプリ本体
# ============================================================
class DartsApp:
    def __init__(self, root):
        self.root = root
        self.root.resizable(False, False)

        self.lang          = "ja"
        self.subject_name  = tk.StringVar(value="")
        self.current_set   = 1
        self.current_throw = 0
        self.throws        = {1: [], 2: []}   # (x_mm, y_mm, x_px, y_px, is_miss)
        self.edit_mode     = None             # None or (set_num, throw_idx)

        self._build_ui()
        self._apply_language()

    # ------------------------------------------------------------------
    # UI 構築
    # ------------------------------------------------------------------
    def _build_ui(self):
        # ヘッダー
        hf = tk.Frame(self.root, bg="#2c3e50", pady=8)
        hf.pack(fill=tk.X)
        self.title_lbl = tk.Label(hf, bg="#2c3e50", fg="white",
                                   font=("Helvetica", 17, "bold"))
        self.title_lbl.pack(side=tk.LEFT, padx=12)
        self.lang_btn = BlueBtn(hf, font=("Helvetica", 10, "bold"),
                                 command=self._toggle_language)
        self.lang_btn.pack(side=tk.RIGHT, padx=12, pady=4)

        # 入力行
        inf = tk.Frame(self.root, padx=12, pady=6)
        inf.pack(fill=tk.X)
        self.name_lbl = tk.Label(inf, font=("Helvetica", 12))
        self.name_lbl.pack(side=tk.LEFT)
        tk.Entry(inf, textvariable=self.subject_name,
                 font=("Helvetica", 12), width=20).pack(side=tk.LEFT, padx=4)
        self.reset_btn = BlueBtn(inf, font=("Helvetica", 11), command=self._reset)
        self.reset_btn.pack(side=tk.RIGHT)

        # メインエリア（キャンバス + サイドパネル）
        mf = tk.Frame(self.root)
        mf.pack(fill=tk.BOTH, padx=8, pady=4)

        self.canvas = tk.Canvas(mf, width=CANVAS_SIZE, height=CANVAS_SIZE,
                                 bg="#1a1a2e", cursor="crosshair",
                                 highlightthickness=0)
        self.canvas.pack(side=tk.LEFT)
        self.canvas.bind("<Button-1>", self._on_canvas_click)

        # サイドパネル
        side = tk.Frame(mf, padx=8)
        side.pack(side=tk.LEFT, fill=tk.Y)

        # ステータス
        self.stat_fr = tk.LabelFrame(side, font=("Helvetica", 11, "bold"),
                                      padx=6, pady=6)
        self.stat_fr.pack(fill=tk.X, pady=(0, 8))
        self.set_lbl = tk.Label(self.stat_fr, font=("Helvetica", 13, "bold"),
                                 fg="#3498db")
        self.set_lbl.pack()
        self.throw_lbl = tk.Label(self.stat_fr, font=("Helvetica", 10),
                                   wraplength=185)
        self.throw_lbl.pack(pady=3)
        self.instr_lbl = tk.Label(self.stat_fr, font=("Helvetica", 9),
                                   fg="#7f8c8d", wraplength=185)
        self.instr_lbl.pack()

        # セット記録
        self.s1_fr = tk.LabelFrame(side, font=("Helvetica", 10, "bold"),
                                    padx=6, pady=4)
        self.s1_fr.pack(fill=tk.X, pady=(0, 4))
        self.set1_labels = [self._make_throw_lbl(self.s1_fr) for _ in range(3)]

        self.s2_fr = tk.LabelFrame(side, font=("Helvetica", 10, "bold"),
                                    padx=6, pady=4)
        self.s2_fr.pack(fill=tk.X, pady=(0, 4))
        self.set2_labels = [self._make_throw_lbl(self.s2_fr) for _ in range(3)]

        # 結果（fill=BOTHで縦に伸びるよう expand=True）
        self.res_fr = tk.LabelFrame(side, font=("Helvetica", 11, "bold"),
                                     padx=6, pady=6)
        self.res_fr.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        self.result_text = tk.Text(self.res_fr, width=24, height=14,
                                    font=("Courier", 10), state=tk.DISABLED,
                                    bg="#f0f0f0", relief=tk.FLAT)
        self.result_text.pack(fill=tk.BOTH, expand=True)

        # ボタン（Excelのみ）
        bf = tk.Frame(side)
        bf.pack(fill=tk.X)

        self.save_btn = BlueBtn(bf, font=("Helvetica", 11),
                                 command=self._save_excel)
        self.save_btn.pack(fill=tk.X, pady=3)
        self.save_btn.set_state("disabled")

        self._draw_dartboard()

    def _make_throw_lbl(self, parent):
        lbl = tk.Label(parent, font=("Helvetica", 10), anchor="w",
                        padx=2, bg="#f5f5f5")
        lbl.pack(fill=tk.X, pady=1)
        return lbl

    # ------------------------------------------------------------------
    # 言語切り替え
    # ------------------------------------------------------------------
    def _toggle_language(self):
        self.lang = "en" if self.lang == "ja" else "ja"
        self._apply_language()

    def _apply_language(self):
        s = STRINGS[self.lang]
        self.root.title(s["title"])
        self.title_lbl.config(text=s["title"])
        self.lang_btn.config(text=s["lang_btn"])
        self.name_lbl.config(text=s["name_label"])
        self.reset_btn.config(text=s["reset_btn"])
        self.stat_fr.config(text=s["status_frame"])
        self.s1_fr.config(text=s["set1_frame"])
        self.s2_fr.config(text=s["set2_frame"])
        self.res_fr.config(text=s["result_frame"])
        self.save_btn.config(text=s["save_btn"])
        self._update_status()
        self._update_throw_labels()
        self.canvas.delete("canvas_hint")
        self.canvas.create_text(CX, CANVAS_SIZE - 8, text=s["canvas_hint"],
                                 fill="#888", font=("Helvetica", 9),
                                 tags="canvas_hint")

    def _t(self, key, **kw):
        t = STRINGS[self.lang][key]
        return t.format(**kw) if kw else t

    # ------------------------------------------------------------------
    # ダーツ盤描画（20分割セグメント詳細版）
    # ------------------------------------------------------------------
    @staticmethod
    def _annular_pts(cx, cy, r_in, r_out, seg_i, steps=16):
        """環状扇形のポリゴン座標"""
        a0 = -90 - 9 + seg_i * 18
        a1 = a0 + 18
        pts = []
        for k in range(steps + 1):
            a = math.radians(a0 + (a1 - a0) * k / steps)
            pts += [cx + r_out * math.cos(a), cy + r_out * math.sin(a)]
        for k in range(steps, -1, -1):
            a = math.radians(a0 + (a1 - a0) * k / steps)
            pts += [cx + r_in * math.cos(a), cy + r_in * math.sin(a)]
        return pts

    def _draw_dartboard(self):
        c = self.canvas
        cx = cy = CX

        # --- 背景 ---
        c.create_rectangle(0, 0, CANVAS_SIZE, CANVAS_SIZE,
                            fill="#1a1a2e", outline="")

        # --- 外縁ブラックリング ---
        c.create_oval(cx - R_SURROUND, cy - R_SURROUND,
                      cx + R_SURROUND, cy + R_SURROUND,
                      fill=SURROUND_C, outline="#3a3a3a", width=2)

        # --- 採点エリアのベース（全体を黒で塗りつぶし） ---
        c.create_oval(cx - R_DBL_OUT, cy - R_DBL_OUT,
                      cx + R_DBL_OUT, cy + R_DBL_OUT,
                      fill=SEG_DARK, outline="")

        # --- 各ゾーンのセグメント（20分割） ---
        # (内径, 外径, 偶数色, 奇数色)
        zones = [
            (R_TRIPLE_OUT, R_DBL_IN,    SEG_DARK,  SEG_LIGHT),   # シングル外
            (R_DBL_IN,     R_DBL_OUT,   RING_RED,  RING_BLUE),   # ダブルリング
            (R_TRIPLE_IN,  R_TRIPLE_OUT, RING_RED, RING_BLUE),   # トリプルリング
            (R_OUTER_BULL, R_TRIPLE_IN, SEG_DARK,  SEG_LIGHT),   # シングル内
        ]
        for (ri, ro, ca, cb) in zones:
            for i in range(20):
                pts = self._annular_pts(cx, cy, ri, ro, i)
                c.create_polygon(pts, fill=(ca if i % 2 == 0 else cb), outline="")

        # --- アウターブル & ブルズアイ ---
        c.create_oval(cx - R_OUTER_BULL, cy - R_OUTER_BULL,
                      cx + R_OUTER_BULL, cy + R_OUTER_BULL,
                      fill=BULL_GREEN, outline=WIRE_C, width=1)
        c.create_oval(cx - R_BULL_EYE, cy - R_BULL_EYE,
                      cx + R_BULL_EYE, cy + R_BULL_EYE,
                      fill=RING_RED, outline=WIRE_C, width=1)

        # --- ワイヤー（放射線） ---
        for i in range(20):
            a = math.radians(-90 - 9 + i * 18)
            c.create_line(cx + R_OUTER_BULL * math.cos(a),
                          cy + R_OUTER_BULL * math.sin(a),
                          cx + R_DBL_OUT * math.cos(a),
                          cy + R_DBL_OUT * math.sin(a),
                          fill=WIRE_C, width=1)

        # --- ワイヤー（円形リング） ---
        for r in (R_DBL_IN, R_TRIPLE_OUT, R_TRIPLE_IN, R_OUTER_BULL):
            c.create_oval(cx - r, cy - r, cx + r, cy + r,
                          fill="", outline=WIRE_C, width=1)
        c.create_oval(cx - R_DBL_OUT, cy - R_DBL_OUT,
                      cx + R_DBL_OUT, cy + R_DBL_OUT,
                      fill="", outline="#555", width=2)

        # --- 数字（外縁リング内） ---
        r_num = (R_DBL_OUT + R_SURROUND) / 2
        for i, num in enumerate(NUMBERS):
            a = math.radians(-90 + i * 18)
            c.create_text(cx + r_num * math.cos(a), cy + r_num * math.sin(a),
                          text=str(num), fill="white",
                          font=("Helvetica", 13, "bold"))

        # --- ヒントテキスト ---
        c.create_text(cx, CANVAS_SIZE - 8, text=self._t("canvas_hint"),
                      fill="#888", font=("Helvetica", 9), tags="canvas_hint")

    # ------------------------------------------------------------------
    # キャンバスクリック & 投擲記録
    # ------------------------------------------------------------------
    def _on_canvas_click(self, event):
        if self.edit_mode is not None:
            sn, ti = self.edit_mode
            self.edit_mode = None
            self._do_throw(event.x, event.y, sn, ti, is_edit=True)
            return
        if self.current_throw >= 3:
            return
        self._do_throw(event.x, event.y, self.current_set,
                       self.current_throw, is_edit=False)

    def _do_throw(self, xp, yp, set_num, throw_idx, is_edit):
        dist_px = math.sqrt((xp - CX) ** 2 + (yp - CY) ** 2)
        is_miss = dist_px > R_DBL_OUT

        if is_miss:
            xm, ym = float(MISS_DISTANCE_MM), 0.0
        else:
            xm = (xp - CX) * PX_TO_MM
            ym = (CY - yp) * PX_TO_MM

        data = (xm, ym, xp, yp, is_miss)

        if is_edit:
            self.throws[set_num][throw_idx] = data
        else:
            self.throws[set_num].append(data)
            self.current_throw += 1

        self._full_redraw()
        self._update_throw_labels()
        self._update_status()

        if not is_edit and self.current_throw == 3:
            if set_num == 1:
                self._complete_set1()
            else:
                self._complete_set2()
        elif is_edit and len(self.throws[1]) == 3 and len(self.throws[2]) == 3:
            result = self._calculate_result()
            self._show_result(result)
            self.save_btn.set_state("normal")

    # ------------------------------------------------------------------
    # マーカー描画
    # ------------------------------------------------------------------
    def _draw_marker(self, xp, yp, num, color):
        r = 9
        self.canvas.create_oval(xp - r, yp - r, xp + r, yp + r,
                                 fill=color, outline="white", width=2)
        self.canvas.create_text(xp, yp, text=str(num),
                                 fill="white", font=("Helvetica", 9, "bold"))

    def _draw_miss_marker(self, seq, throw_num, side):
        """ミスマーカーをキャンバス隅に表示（board外のコーナー領域）"""
        bx = 18 if side == "L" else CANVAS_SIZE - 18
        by = 16 + seq * 28
        anc = "w" if side == "L" else "e"
        self.canvas.create_text(bx, by, text=f"✕{throw_num} MISS",
                                 fill="#ff4444", font=("Helvetica", 10, "bold"),
                                 anchor=anc)

    def _full_redraw(self):
        """キャンバスを全クリアして全データを再描画"""
        self.canvas.delete("all")
        self._draw_dartboard()

        c1, c2 = "#3498db", "#e67e22"
        ms1 = ms2 = 0
        for i, t in enumerate(self.throws[1]):
            if t[4]:
                self._draw_miss_marker(ms1, i + 1, "L")
                ms1 += 1
            else:
                self._draw_marker(t[2], t[3], i + 1, c1)

        for i, t in enumerate(self.throws[2]):
            if t[4]:
                self._draw_miss_marker(ms2, i + 1, "R")
                ms2 += 1
            else:
                self._draw_marker(t[2], t[3], i + 1, c2)

        if len(self.throws[1]) == 3:
            self._draw_grouping_circle(self.throws[1], c1, 1)
        if len(self.throws[2]) == 3:
            self._draw_grouping_circle(self.throws[2], c2, 2)

    def _draw_grouping_circle(self, throws, color, set_num):
        visual = [t for t in throws if not t[4]]
        if not visual:
            return
        xs, ys = [t[2] for t in visual], [t[3] for t in visual]
        gcx = sum(xs) / len(xs)
        gcy = sum(ys) / len(ys)
        if len(visual) >= 2:
            rad = max(math.sqrt((x - gcx) ** 2 + (y - gcy) ** 2)
                      for x, y in zip(xs, ys))
            self.canvas.create_oval(gcx - rad, gcy - rad,
                                     gcx + rad, gcy + rad,
                                     outline=color, width=2, dash=(6, 4))
        self.canvas.create_oval(gcx - 5, gcy - 5, gcx + 5, gcy + 5,
                                 fill=color, outline="white", width=2)
        self.canvas.create_text(gcx, gcy - 14,
                                 text=self._t("centroid_label", n=set_num),
                                 fill=color, font=("Helvetica", 8, "bold"))

    # ------------------------------------------------------------------
    # 記録ラベル更新（クリックで修正モード切替）
    # ------------------------------------------------------------------
    def _update_throw_labels(self):
        for sn in (1, 2):
            lbls = self.set1_labels if sn == 1 else self.set2_labels
            for i, lbl in enumerate(lbls):
                recorded = i < len(self.throws[sn])
                in_edit  = (self.edit_mode == (sn, i))

                if recorded:
                    t = self.throws[sn][i]
                    if t[4]:
                        txt = self._t("throw_miss", n=i + 1)
                    else:
                        txt = self._t("throw_recorded", n=i + 1, x=t[0], y=t[1])
                    bg = "#ffe066" if in_edit else "#e8f4ff"
                    fg = "#c0392b" if in_edit else "#2c3e50"
                    lbl.config(text=txt, fg=fg, bg=bg, cursor="hand2")
                    lbl.bind("<Button-1>",
                              lambda e, s=sn, ii=i: self._on_label_click(s, ii))
                    lbl.bind("<Enter>",
                              lambda e, l=lbl, s=sn, ii=i:
                              l.config(bg="#ffe066") if self.edit_mode != (s, ii) else None)
                    lbl.bind("<Leave>",
                              lambda e, l=lbl, s=sn, ii=i:
                              l.config(bg="#ffe066" if self.edit_mode == (s, ii) else "#e8f4ff"))
                else:
                    lbl.config(text=self._t("throw_pending", n=i + 1),
                               fg="#95a5a6", bg="#f5f5f5", cursor="arrow")
                    lbl.unbind("<Button-1>")
                    lbl.unbind("<Enter>")
                    lbl.unbind("<Leave>")

    def _on_label_click(self, sn, throw_idx):
        if self.edit_mode == (sn, throw_idx):
            self.edit_mode = None
        else:
            self.edit_mode = (sn, throw_idx)
        self._update_throw_labels()
        self._update_status()

    # ------------------------------------------------------------------
    # ステータス更新
    # ------------------------------------------------------------------
    def _update_status(self):
        s = STRINGS[self.lang]
        self.set_lbl.config(text=self._t("set_label", n=self.current_set))

        if self.edit_mode is not None:
            sn, ti = self.edit_mode
            self.throw_lbl.config(text=self._t("edit_mode_msg", s=sn, n=ti + 1))
            self.instr_lbl.config(text=self._t("edit_cancel_tip"))
            return

        if self.current_throw < 3:
            self.throw_lbl.config(text=self._t("throw_prompt", n=self.current_throw + 1))
            self.instr_lbl.config(
                text=s["instruction"] if self.current_set == 1 else s["set2_instruction"]
            )
        else:
            key = "set1_done_status" if self.current_set == 1 else "set2_done_status"
            self.throw_lbl.config(text=s[key])
            self.instr_lbl.config(text="")

    # ------------------------------------------------------------------
    # セット完了
    # ------------------------------------------------------------------
    def _complete_set1(self):
        # ダイアログなしで即セット2へ自動移行
        self._go_to_set2()

    def _go_to_set2(self):
        self.current_set   = 2
        self.current_throw = 0
        self.set_lbl.config(fg="#e67e22")
        self._update_status()

    def _complete_set2(self):
        result = self._calculate_result()
        self._show_result(result)
        self.save_btn.set_state("normal")

    # ------------------------------------------------------------------
    # 計算 & 結果表示
    # ------------------------------------------------------------------
    def _calculate_result(self):
        def grouping_r(throws):
            xs, ys = [t[0] for t in throws], [t[1] for t in throws]
            gx = sum(xs) / len(xs)
            gy = sum(ys) / len(ys)
            return max(math.sqrt((x - gx) ** 2 + (y - gy) ** 2)
                       for x, y in zip(xs, ys))

        def center_d(throws):
            return sum(math.sqrt(t[0] ** 2 + t[1] ** 2) for t in throws) / len(throws)

        r1 = grouping_r(self.throws[1]);  r2 = grouping_r(self.throws[2])
        d1 = center_d(self.throws[1]);    d2 = center_d(self.throws[2])
        imp = ((r1 - r2) / r1 * 100) if r1 > 0 else 0
        acc = ((d1 - d2) / d1 * 100) if d1 > 0 else 0
        return {"radius_1": r1, "radius_2": r2, "improvement": imp,
                "center_dist_1": d1, "center_dist_2": d2,
                "accuracy_improvement": acc, "improved": r2 < r1}

    def _show_result(self, r):
        imp_v  = abs(r["improvement"])
        acc_v  = r["accuracy_improvement"]
        imp_ok = r["improved"]
        acc_ok = acc_v > 0

        if self.lang == "ja":
            verdict = "★ 上達しました！" if imp_ok else "▼ バラつきが増えました"
            lines = [
                "═" * 24,
                "  【バラつき（グルーピング）】",
                f"  1回目：{r['radius_1']:.1f} mm",
                f"  2回目：{r['radius_2']:.1f} mm",
                f"  改善率：{imp_v:.0f}% {'↑ 改善' if imp_ok else '↓ 悪化'}",
                "",
                "  【正確さ（中心距離平均）】",
                f"  1回目：{r['center_dist_1']:.1f} mm",
                f"  2回目：{r['center_dist_2']:.1f} mm",
                f"  改善率：{acc_v:.0f}% {'↑ 改善' if acc_ok else '↓ 悪化'}",
                "",
                "─" * 24,
                f"  {verdict}",
                "─" * 24,
            ]
        else:
            verdict = "★ Improved!" if imp_ok else "▼ Spread increased"
            lines = [
                "═" * 24,
                "  [Grouping / Spread]",
                f"  Round 1: {r['radius_1']:.1f} mm",
                f"  Round 2: {r['radius_2']:.1f} mm",
                f"  Change:  {imp_v:.0f}% {'↑ Better' if imp_ok else '↓ Worse'}",
                "",
                "  [Accuracy (avg from bull)]",
                f"  Round 1: {r['center_dist_1']:.1f} mm",
                f"  Round 2: {r['center_dist_2']:.1f} mm",
                f"  Change:  {acc_v:.0f}% {'↑ Better' if acc_ok else '↓ Worse'}",
                "",
                "─" * 24,
                f"  {verdict}",
                "─" * 24,
            ]

        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete("1.0", tk.END)
        self.result_text.insert(tk.END, "\n".join(lines))
        self.result_text.config(state=tk.DISABLED)

    # ------------------------------------------------------------------
    # リセット
    # ------------------------------------------------------------------
    def _do_reset(self):
        self.throws        = {1: [], 2: []}
        self.current_set   = 1
        self.current_throw = 0
        self.edit_mode     = None
        self.set_lbl.config(fg="#3498db")
        self.save_btn.set_state("disabled")
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete("1.0", tk.END)
        self.result_text.config(state=tk.DISABLED)
        self.canvas.delete("all")
        self._draw_dartboard()
        self._update_status()
        self._update_throw_labels()

    def _reset(self):
        if messagebox.askyesno(self._t("reset_confirm_title"),
                               self._t("reset_confirm_msg")):
            self._do_reset()

    # ------------------------------------------------------------------
    # Excel 出力（保存後に自動リセット）
    # ------------------------------------------------------------------
    def _save_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror(self._t("no_excel_title"), self._t("no_excel_msg"))
            return

        name = self.subject_name.get().strip() or ("体験者" if self.lang == "ja" else "Player")
        now  = datetime.now()
        default_fn = f"darts_{name}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[(self._t("excel_filetypes"), "*.xlsx")],
            initialfile=default_fn
        )
        if not fp:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._t("xl_sheet")

        hf  = PatternFill("solid", fgColor="2C3E50")
        s1f = PatternFill("solid", fgColor="D6EAF8")
        s2f = PatternFill("solid", fgColor="FDEBD0")
        rsf = PatternFill("solid", fgColor="D5F5E3")
        bold = Font(bold=True)
        cnt  = Alignment(horizontal="center", vertical="center")
        thin = Border(left=Side(style="thin"),  right=Side(style="thin"),
                      top=Side(style="thin"),   bottom=Side(style="thin"))

        def wc(row, col, val, fill=None, font=None, align=None):
            cell = ws.cell(row=row, column=col, value=val)
            if fill:  cell.fill  = fill
            if font:  cell.font  = font
            if align: cell.alignment = align
            cell.border = thin
            return cell

        # タイトル & 基本情報
        ws.merge_cells("A1:G1")
        tc = ws.cell(row=1, column=1, value=self._t("xl_main_title"))
        tc.font = Font(bold=True, size=16, color="FFFFFF")
        tc.fill = hf; tc.alignment = cnt
        wc(2, 1, self._t("xl_name"), font=bold); wc(2, 2, name)
        wc(2, 3, self._t("xl_date"), font=bold); wc(2, 4, now.strftime("%Y/%m/%d %H:%M:%S"))
        ws.append([])

        def write_set(row0, throws, sfill, hcolor, set_key):
            ws.merge_cells(f"A{row0}:G{row0}")
            h = ws.cell(row=row0, column=1, value=self._t(set_key))
            h.font = Font(bold=True, size=12, color=hcolor)
            h.fill = sfill; h.border = thin
            cf = PatternFill("solid", fgColor="AED6F1" if sfill == s1f else "FAD7A0")
            for ci, txt in enumerate([self._t("xl_throw"), self._t("xl_x"),
                                       self._t("xl_y"),    self._t("xl_dist")], 1):
                wc(row0 + 1, ci, txt, fill=cf, font=bold, align=cnt)
            for i, t in enumerate(throws):
                xm, ym, _, _, miss = t
                dist = math.sqrt(xm ** 2 + ym ** 2)
                lbl = self._t("xl_miss") if miss else self._t("xl_throw_n", n=i + 1)
                wc(row0 + 2 + i, 1, lbl, fill=sfill)
                wc(row0 + 2 + i, 2, "---" if miss else round(xm, 2))
                wc(row0 + 2 + i, 3, "---" if miss else round(ym, 2))
                wc(row0 + 2 + i, 4, round(dist, 2))
            xms = [t[0] for t in throws]; yms = [t[1] for t in throws]
            gcx = sum(xms) / len(xms);    gcy = sum(yms) / len(yms)
            gr  = max(math.sqrt((x - gcx) ** 2 + (y - gcy) ** 2)
                      for x, y in zip(xms, yms))
            wc(row0 + 5, 1, self._t("xl_group_r"), font=bold, fill=sfill)
            wc(row0 + 5, 2, round(gr, 2)); wc(row0 + 5, 3, self._t("xl_mm"))
            return gr

        r1 = write_set(4,  self.throws[1], s1f, "1A5276", "xl_set1")
        r2 = write_set(11, self.throws[2], s2f, "7D3C98", "xl_set2")

        imp  = (r1 - r2) / r1 * 100 if r1 > 0 else 0
        d1a  = sum(math.sqrt(t[0] ** 2 + t[1] ** 2) for t in self.throws[1]) / 3
        d2a  = sum(math.sqrt(t[0] ** 2 + t[1] ** 2) for t in self.throws[2]) / 3
        accp = (d1a - d2a) / d1a * 100 if d1a > 0 else 0

        ws.merge_cells("A18:G18")
        hr = ws.cell(row=18, column=1, value=self._t("xl_result"))
        hr.font = Font(bold=True, size=13, color="1E8449")
        hr.fill = rsf; hr.border = thin

        hrf = PatternFill("solid", fgColor="A9DFBF")
        result_rows = [
            (self._t("xl_item"),   self._t("xl_round1"), self._t("xl_round2"),
             self._t("xl_improve"), self._t("xl_judge")),
            (self._t("xl_spread"),
             f"{r1:.1f} mm", f"{r2:.1f} mm", f"{imp:.0f}%",
             self._t("xl_good") if r2 < r1 else self._t("xl_bad")),
            (self._t("xl_accuracy"),
             f"{d1a:.1f} mm", f"{d2a:.1f} mm", f"{accp:.0f}%",
             self._t("xl_good") if d2a < d1a else self._t("xl_bad")),
        ]
        for ri, rd in enumerate(result_rows):
            for ci, v in enumerate(rd):
                wc(19 + ri, ci + 1, v,
                   fill=hrf if ri == 0 else rsf,
                   font=bold if ri == 0 else None, align=cnt)

        wc(23, 1, self._t("xl_comment"), font=bold)
        ws.merge_cells("B23:G23"); ws.cell(row=23, column=2).border = thin

        for i, w in enumerate([22, 15, 15, 22, 15, 10, 10], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for rn, rh in {1: 30, 4: 22, 11: 22, 18: 24}.items():
            ws.row_dimensions[rn].height = rh

        wb.save(fp)
        messagebox.showinfo(self._t("save_done_title"),
                            self._t("save_done_msg", path=fp))
        # Excel 保存後に自動リセット
        self._do_reset()


# ============================================================
# エントリーポイント
# ============================================================
def main():
    root = tk.Tk()
    root.configure(bg="#ecf0f1")
    DartsApp(root)

    root.update_idletasks()
    w = root.winfo_width()
    h = root.winfo_height()
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.mainloop()


if __name__ == "__main__":
    main()
